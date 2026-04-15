# ================= IMPORTAÇÕES =================

from flask import (
    Flask,
    render_template,
    jsonify,
    request,
    redirect,
    url_for,
    session,
    flash,
    send_file,
    Response,
    stream_with_context,
)

import subprocess
import io
import threading
import re

from openpyxl import load_workbook, Workbook
from functools import wraps
import os
import calendar
import datetime
import time
import yt_dlp
import uuid


# ================= CACHE =================

cache_total_geral = {"dados": None, "tempo": 0}
TEMPO_CACHE = 300  # 5 minutos


# ================= CONFIG FLASK =================

app = Flask(__name__, static_folder="static", template_folder="templates")
app.secret_key = "NWanClh3BDY8I67SwHmXjhPQ2We2n2GMbr7KOtRIeJ7s9KMOMp"


# 🔒 CONFIGURAÇÕES PARA NÃO SALVAR LOGIN

app.config.update(
    SESSION_PERMANENT=False,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=False,
)


@app.after_request
def no_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


# ================= USUÁRIOS =================

USUARIOS = {"admin": "sig@2025"}


# ================= CONFIG EXCEL =================

ARQUIVO_SIG = "dados.xlsx"
ARQUIVO_SSH = "dadossh.xlsx"
ANO_FIXO = 2026

# ================= MESES GLOBAIS =================
MESES_VALIDOS = [
    "JANEIRO",
    "FEVEREIRO",
    "MARÇO",
    "ABRIL",
    "MAIO",
    "JUNHO",
    "JULHO",
    "AGOSTO",
    "SETEMBRO",
    "OUTUBRO",
    "NOVEMBRO",
    "DEZEMBRO",
]


# ================= LOGIN =================


@app.route("/Login-Planilha", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        session.clear()

        usuario = request.form.get("usuario")
        senha = request.form.get("senha")

        if usuario in USUARIOS and USUARIOS[usuario] == senha:
            session["usuario"] = usuario
            session.permanent = False
            return redirect(url_for("planilha_sig"))

        flash("Usuário ou senha incorretos")

    return render_template("login.html")


# ================= MIDDLEWARE =================


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "usuario" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated_function


# ================= HELPERS =================


def garantir_arquivo(arquivo):
    if not os.path.exists(arquivo):
        wb = Workbook()
        wb.save(arquivo)
        wb.close()


def garantir_aba(arquivo, mes, tipo):
    garantir_arquivo(arquivo)
    mes = mes.upper()

    wb = load_workbook(arquivo)

    if mes == "TOTAL GERAL":
        wb.close()
        return

    if mes not in wb.sheetnames:
        ws = wb.create_sheet(mes)

        ws["A1"] = "ID"
        ws["B1"] = "DATA"
        ws["C1"] = "P&R"

        if tipo == "sig":
            ws["D1"] = "EMBAIXADOR"
            ws["F1"] = "CSS"
            ws["G1"] = "% CSS"

        meses = {
            "JANEIRO": 1,
            "FEVEREIRO": 2,
            "MARÇO": 3,
            "ABRIL": 4,
            "MAIO": 5,
            "JUNHO": 6,
            "JULHO": 7,
            "AGOSTO": 8,
            "SETEMBRO": 9,
            "OUTUBRO": 10,
            "NOVEMBRO": 11,
            "DEZEMBRO": 12,
        }

        numero = meses.get(mes, 1)
        ultimo = calendar.monthrange(ANO_FIXO, numero)[1]

        for d in range(1, ultimo + 1):
            data = datetime.date(ANO_FIXO, numero, d)
            ws.cell(row=d + 1, column=1, value=d)
            ws.cell(row=d + 1, column=2, value=data.strftime("%d/%m/%Y"))

        wb.save(arquivo)

    wb.close()


def corrigir_zeros(arquivo):
    wb = load_workbook(arquivo)

    for aba in wb.sheetnames:
        ws = wb[aba]

        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=3).value in (None, ""):
                ws.cell(row=r, column=3, value=0)

            if ws.cell(row=r, column=4).value in (None, ""):
                ws.cell(row=r, column=4, value=0)

            if ws.cell(row=r, column=6).value in (None, ""):
                ws.cell(row=r, column=6, value=0)

            if ws.cell(row=r, column=7).value in (None, ""):
                ws.cell(row=r, column=7, value=0)

    wb.save(arquivo)
    wb.close()


def soma_coluna(arquivo, coluna):
    """Soma apenas abas dos 12 meses com dados (max_row > 1)"""
    if not os.path.exists(arquivo):
        print(f"soma_coluna: {arquivo} não existe")
        return 0

    MESES_ORDEM = [
        "JANEIRO",
        "FEVEREIRO",
        "MARÇO",
        "ABRIL",
        "MAIO",
        "JUNHO",
        "JULHO",
        "AGOSTO",
        "SETEMBRO",
        "OUTUBRO",
        "NOVEMBRO",
        "DEZEMBRO",
    ]

    try:
        wb = load_workbook(arquivo, data_only=True)
        total = 0.0
        abas_com_dados = []

        for mes in MESES_ORDEM:
            if mes in wb.sheetnames:
                ws = wb[mes]
                if ws.max_row > 1:  # Tem dados além do header
                    abas_com_dados.append(mes)
                    for row in ws.iter_rows(
                        min_row=2, max_col=coluna, values_only=True
                    ):
                        if len(row) >= coluna and row[coluna - 1] is not None:
                            try:
                                val = float(str(row[coluna - 1]).replace(",", "."))
                                total += abs(val)  # Soma valor absoluto
                            except (ValueError, TypeError):
                                pass

        print(
            f"soma_coluna({arquivo}, col{coluna}): abas_com_dados={abas_com_dados}, total={total}"
        )
        wb.close()
        return total
    except Exception as e:
        print(f"ERRO soma_coluna({arquivo}): {e}")
        return 0


def garantir_total_geral(arquivo):
    """Cria/atualiza aba TOTAL GERAL com somas das mensais"""
    if not os.path.exists(arquivo):
        return

    wb = load_workbook(arquivo)
    if "TOTAL GERAL" not in wb.sheetnames:
        wb.create_sheet("TOTAL GERAL")

    # Recalcula somas se necessário (simplificado)
    ws = wb["TOTAL GERAL"]
    ws["C2"] = soma_coluna(arquivo, 3)  # P&R total
    ws["F2"] = soma_coluna(arquivo, 6)  # CSS total

    wb.save(arquivo)
    wb.close()


# ================= ROTAS =================


@app.route("/Home")
@login_required
def home():
    return render_template("inicio.html")


@app.route("/planilha-sig")
@login_required
def planilha_sig():
    return render_template("index.html", tipo="sig")


@app.route("/planilha-ssh")
@login_required
def planilha_ssh():
    return render_template("index.html", tipo="ssh")


# ================= API MESES =================


@app.route("/api/meses")
@login_required
def api_meses():
    garantir_arquivo(ARQUIVO_SIG)

    wb = load_workbook(ARQUIVO_SIG, read_only=True)

    MESES_ORDEM = [
        "JANEIRO",
        "FEVEREIRO",
        "MARÇO",
        "ABRIL",
        "MAIO",
        "JUNHO",
        "JULHO",
        "AGOSTO",
        "SETEMBRO",
        "OUTUBRO",
        "NOVEMBRO",
        "DEZEMBRO",
        "TOTAL GERAL",
    ]

    abas = set(s.strip().upper() for s in wb.sheetnames)
    wb.close()

    resultado = [mes for mes in MESES_ORDEM if mes in abas]

    return jsonify(resultado)


# ================= API DIAS =================


@app.route("/api/dias")
@login_required
def api_dias():
    mes = request.args.get("mes")
    if not mes:
        return jsonify([])

    meses = {
        "JANEIRO": 1,
        "FEVEREIRO": 2,
        "MARÇO": 3,
        "ABRIL": 4,
        "MAIO": 5,
        "JUNHO": 6,
        "JULHO": 7,
        "AGOSTO": 8,
        "SETEMBRO": 9,
        "OUTUBRO": 10,
        "NOVEMBRO": 11,
        "DEZEMBRO": 12,
    }

    numero = meses.get(mes.upper(), 1)
    ultimo = calendar.monthrange(ANO_FIXO, numero)[1]
    return jsonify(list(range(1, ultimo + 1)))


# ================= API SALVAR =================


@app.route("/api/salvar", methods=["POST"])
@login_required
def api_salvar():
    try:
        data = request.json
        mes = data.get("mes")
        dia = int(data.get("dia"))
        pr = data.get("pr")
        emb = data.get("emb")
        css = data.get("css")
        tipo = data.get("tipo")
        percent_css = data.get("percent_css")

        if mes and mes.upper() == "TOTAL GERAL":
            return jsonify({"error": "TOTAL GERAL não pode ser editado"}), 403

        arquivo = ARQUIVO_SIG if tipo == "sig" else ARQUIVO_SSH
        garantir_aba(arquivo, mes, tipo)

        wb = load_workbook(arquivo)
        ws = wb[mes.upper()]

        if pr not in (None, ""):
            ws.cell(row=dia + 1, column=3, value=float(str(pr).replace(",", ".")))

        ws.cell(row=dia + 1, column=4, value=emb if emb else "")

        if css not in (None, ""):
            ws.cell(row=dia + 1, column=6, value=float(str(css).replace(",", ".")))

        if percent_css not in (None, ""):
            ws.cell(
                row=dia + 1, column=7, value=float(str(percent_css).replace(",", "."))
            )

        wb.save(arquivo)
        wb.close()

        # Limpa cache ao salvar
        global cache_total_geral
        cache_total_geral["dados"] = None

        return jsonify({"ok": True})

    except Exception as e:
        print("ERRO AO SALVAR:", str(e))
        return jsonify({"error": str(e)}), 500


# ================= API TABELA =================


@app.route("/api/tabela")
@login_required
def api_tabela():
    mes = request.args.get("mes")
    tipo = request.args.get("tipo")

    arquivo = ARQUIVO_SIG if tipo == "sig" else ARQUIVO_SSH
    if not os.path.exists(arquivo):
        return jsonify([])

    wb = load_workbook(arquivo, data_only=True)
    if mes.upper() not in wb.sheetnames:
        wb.close()
        return jsonify([])

    ws = wb[mes.upper()]
    dados = []

    for r in range(2, ws.max_row + 1):
        data_celula = ws.cell(row=r, column=2).value

        if isinstance(data_celula, (datetime.date, datetime.datetime)):
            data_formatada = data_celula.strftime("%d/%m/%Y")
        else:
            data_formatada = data_celula or ""

        item = {
            "id": ws.cell(row=r, column=1).value or "",
            "data": data_formatada,
            "pr": ws.cell(row=r, column=3).value or "",
        }

        if tipo == "sig":
            item["emb"] = ws.cell(row=r, column=4).value or ""
            item["css"] = ws.cell(row=r, column=6).value or ""
            item["percent_css"] = ws.cell(row=r, column=7).value or ""

        dados.append(item)

    wb.close()
    return jsonify(dados)


# ================= RESUMO =================


@app.route("/resumo")
@login_required
def resumo():
    # Garante arquivos e TOTAL GERAL
    garantir_arquivo(ARQUIVO_SIG)
    garantir_arquivo(ARQUIVO_SSH)
    garantir_total_geral(ARQUIVO_SIG)
    garantir_total_geral(ARQUIVO_SSH)

    try:
        total_sig_pr = soma_coluna(ARQUIVO_SIG, 3)
        total_ssh_pr = soma_coluna(ARQUIVO_SSH, 3)
        total_sig_css = soma_coluna(ARQUIVO_SIG, 6)

        resultado = int(total_sig_pr or 0) - int(total_ssh_pr or 0)

        print(
            f"SIG P&R detalhado: arquivos={os.path.exists(ARQUIVO_SIG)}, total={total_sig_pr}"
        )
        print(f"SIG CSS detalhado: total={total_sig_css}")
        print(
            f"SSH P&R detalhado: arquivos={os.path.exists(ARQUIVO_SSH)}, total={total_ssh_pr}"
        )
        print(f"RESULTADO final: {resultado}")

        return render_template(
            "resumo.html",
            total_sig=int(total_sig_pr or 0),
            total_ssh=int(total_ssh_pr or 0),
            total_css=int(total_sig_css or 0),
            resultado=resultado,
        )
    except Exception as e:
        print(f"ERRO /resumo: {str(e)}")
    import traceback

    print(f"ERRO DETALHADO /resumo: {str(e)}")
    print(traceback.format_exc())
    flash(f"Erro no resumo: {str(e)}")
    return (
        render_template(
            "resumo.html", total_sig=0, total_ssh=0, total_css=0, resultado=0
        ),
        500,
    )


# ================= API TOTAL GERAL =================


@app.route("/api/mes-total-geral")
@login_required
def api_mes_total_geral():
    global cache_total_geral

    tipo = request.args.get("tipo")
    arquivo = ARQUIVO_SIG if tipo == "sig" else ARQUIVO_SSH

    # ✅ Verifica cache com tempo
    tempo_atual = time.time()
    if (
        cache_total_geral["dados"] is not None
        and (tempo_atual - cache_total_geral["tempo"]) < TEMPO_CACHE
    ):
        return jsonify(cache_total_geral["dados"])

    if not os.path.exists(arquivo):
        return jsonify([])

    MESES_ORDEM = [
        "JANEIRO",
        "FEVEREIRO",
        "MARÇO",
        "ABRIL",
        "MAIO",
        "JUNHO",
        "JULHO",
        "AGOSTO",
        "SETEMBRO",
        "OUTUBRO",
        "NOVEMBRO",
        "DEZEMBRO",
    ]

    wb = load_workbook(arquivo, read_only=True, data_only=True)
    resultado = []

    total_pr_anual = 0
    soma_css_peso_anual = 0
    soma_css_anual = 0

    soma_css_anual = 0
    soma_css_peso_anual = 0

    for mes in MESES_ORDEM:
        if mes not in wb.sheetnames:
            continue

        ws = wb[mes]

        total_pr_mes = 0
        soma_css_mes = 0
        soma_css_peso_mes = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 7:
                continue
            pr = row[2]
            css = row[5]
            percent = row[6]

            if pr is not None and str(pr).strip():
                try:
                    total_pr_mes += float(str(pr).replace(",", "."))
                except:
                    pass

            if (
                css is not None
                and percent is not None
                and str(css).strip()
                and str(percent).strip()
            ):
                try:
                    css_val = float(str(css).replace(",", "."))
                    percent_val = float(str(percent).replace(",", "."))
                    if css_val > 0:
                        soma_css_mes += css_val
                        soma_css_peso_mes += css_val * percent_val
                except:
                    pass

        media_percent_mes = (
            round(soma_css_peso_mes / soma_css_mes, 1) if soma_css_mes > 0 else 0
        )

        resultado.append(
            {
                "id": "",
                "data": mes,
                "pr": int(total_pr_mes),
                "css": int(soma_css_mes),
                "percent_css": media_percent_mes,
            }
        )

        total_pr_anual += total_pr_mes
        soma_css_anual += soma_css_mes
        soma_css_peso_anual += soma_css_peso_mes

    wb.close()

    media_percent_anual = (
        round(soma_css_peso_anual / soma_css_anual, 1) if soma_css_anual > 0 else 0
    )

    resultado.append(
        {
            "id": "",
            "data": "TOTAL GERAL",
            "pr": int(total_pr_anual),
            "css": int(soma_css_anual),
            "percent_css": media_percent_anual,
        }
    )

    # ✅ Salva cache com timestamp
    cache_total_geral["dados"] = resultado
    cache_total_geral["tempo"] = tempo_atual

    return jsonify(resultado)


# ================= DOWNLOAD (PERFEITO SEM LOGIN) =================
PASTA_DOWNLOAD = "downloads"
os.makedirs(PASTA_DOWNLOAD, exist_ok=True)


@app.route("/api/info", methods=["POST"])
def info_video():
    try:
        data = request.get_json()
        url = data.get("url")
        if not url:
            return jsonify({"erro": "URL vazia"})

        # ✅ Config otimizada para TODAS plataformas
        ydl_opts = {
            "quiet": True,
            "skip_download": True,
            "extract_flat": False,  # Pega info completa
        }

        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(url, download=False)

            duracao = info.get("duration", 0) or 0
            filesize = info.get("filesize", 0) or info.get("filesize_approx", 0)

            return jsonify(
                {
                    "titulo": info.get("title", "Sem título")[:100],
                    "duracao": f"{int(duracao//60):02d}:{int(duracao%60):02d}",
                    "tamanho_video": (
                        f"{filesize/1024/1024:.1f} MB" if filesize else "N/D"
                    ),
                    "tamanho_audio": (
                        f"{filesize/10/1024/1024:.1f} MB" if filesize else "N/D"
                    ),
                    "thumbnail": info.get("thumbnail")
                    or info.get("thumbnails", [{}])[0].get("url"),
                }
            )
    except Exception as e:
        print("ERRO INFO:", e)
        return jsonify({"erro": str(e)}), 500


@app.route("/api/download", methods=["POST"])
def api_download():
    try:
        data = request.get_json()
        url, tipo = data.get("url"), data.get("tipo", "video")
        if not url:
            return jsonify({"erro": "URL obrigatória"}), 400

        nome_uuid = str(uuid.uuid4())
        extensao = "mp3" if tipo == "audio" else "mp4"
        caminho = os.path.join(PASTA_DOWNLOAD, f"{nome_uuid}.{extensao}")

        # ✅ CMD PERFEITO para todas plataformas
        cmd_base = [
            "python",
            "-m",
            "yt_dlp",
            "--no-warnings",  # Menos spam
            "--embed-subs",  # Legendas se tiver
            "--embed-thumbnail",  # Thumbnail no vídeo
            "-o",
            caminho,
        ]

        if tipo == "audio":
            cmd = cmd_base + [
                "-x",
                "--audio-format",
                "mp3",
                "--audio-quality",
                "192K",
                url,
            ]
        else:
            cmd = cmd_base + [
                "-f",
                "best[ext=mp4][height<=720]/best[height<=720]/best",
                url,
            ]

        print("Executando:", " ".join(cmd))  # Debug

        result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)

        if result.returncode != 0:
            return jsonify({"erro": result.stderr[:300] or "Falhou"}), 500

        if not os.path.exists(caminho) or os.path.getsize(caminho) < 1000:
            return jsonify({"erro": "Arquivo inválido ou muito pequeno"}), 500

        # Nome bonito
        try:
            with yt_dlp.YoutubeDL({"quiet": True}) as ydl:
                info = ydl.extract_info(url, download=False)
                title = re.sub(r"[^\w\s.-]", "", info.get("title", "video")[:50])
        except:
            title = "video"

        # Cleanup
        def cleanup():
            time.sleep(180)
            if os.path.exists(caminho):
                os.remove(caminho)

        threading.Thread(target=cleanup, daemon=True).start()

        return jsonify(
            {
                "success": True,
                "path": caminho,
                "download_name": f"{title}.{extensao}",
                "tamanho": f"{os.path.getsize(caminho)/1024/1024:.1f} MB",
            }
        )

    except subprocess.TimeoutExpired:
        return jsonify({"erro": "Timeout (vídeo muito grande)"}), 500
    except Exception as e:
        print("ERRO DOWNLOAD:", e)
        return jsonify({"erro": str(e)}), 500


# ================= PAGINA DOWNLOAD =================
@app.route("/download")
def pagina_download():
    return render_template("download.html")


# ================= OUTRAS ROTAS =================


@app.route("/calculadora")
@login_required
def calculadora():
    return render_template("calculadora.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/baixar-sig")
def baixar_sig():
    if os.path.exists(ARQUIVO_SIG):
        return send_file(
            ARQUIVO_SIG,
            as_attachment=True,
            download_name="Planilha_SIG_Atualizada.xlsx",
        )
    else:
        return "Arquivo não encontrado", 404


@app.route("/baixar-ssh")
def baixar_ssh():
    if os.path.exists(ARQUIVO_SSH):
        return send_file(
            ARQUIVO_SSH,
            as_attachment=True,
            download_name="Planilha_SSH_Atualizada.xlsx",
        )
    else:
        return "Arquivo não encontrado", 404


@app.route("/")
def index():
    if "usuario" in session:
        return redirect(url_for("planilha_sig"))
    return redirect(url_for("login"))


# ================= RUN =================

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
